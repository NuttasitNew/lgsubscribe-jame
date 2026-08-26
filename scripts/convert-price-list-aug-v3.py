#!/usr/bin/env python3
"""Convert LG Subscribe Price list_Aug_V3.pdf into a checkable Excel workbook."""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

PDF_PATH = Path("/Users/nuttasit/lg-jame/Price list_Aug_V3.pdf")
OUT_PATH = Path("/Users/nuttasit/lg-jame/Price list_Aug_V3.xlsx")
LAYOUT_TXT = Path("/tmp/pricelist/layout.txt")

CONTRACT_RE = re.compile(
    r"^(?:\d+Y[_\s]?(?:Visit|Self|No\s*Service)?|No\s*Service)$",
    re.I,
)
MODEL_TOKEN_RE = re.compile(
    r"(?:NEW\s+)?"
    r"([A-Z]{1,6}[-]?\d{2,}[A-Z0-9._/-]*"
    r"|A9T-[A-Z]+(?:\.[A-Z0-9]+)?"
    r"|ART\d{2}A(?:\.[A-Z0-9]+)?"
    r"|S(?:AQ|IQ)\d{2}[AB]?(?:\.[A-Z0-9]+)?"
    r"|S95TR|S80TY|S70TY|S30A"
    r"|WD516|WD518|WD110MN"
    r"|PTOL[A-Z0-9.]+"
    r"|PTODFC[A-Z0-9.]+)",
    re.I,
)
MONEY_RE = re.compile(r"^-?\d{1,3}(?:,\d{3})*(?:\.\d+)?$")
PROMO_HEADER_RE = re.compile(
    r"โปรโมช[ัน์]*\s*(\d{1,2}\s*ส\.ค\.?\s*[-–]\s*\d{1,2}\s*ส\.ค\.?\s*2569)",
    re.I,
)
BILL_PRICE_RE = re.compile(
    r"รอบบิลที่\s*([\d\s\-–]+)\s*\(([\d,\.]+)\)",
    re.I,
)


def norm_space(s: str) -> str:
    s = unicodedata.normalize("NFC", s or "")
    s = s.replace("\u00a0", " ").replace("\n", " | ")
    s = re.sub(r"[ \t]+", " ", s)
    s = s.replace(" | | ", " | ").strip(" |")
    return s.strip()


def clean_cell(s) -> str:
    if s is None:
        return ""
    s = norm_space(str(s))
    s = s.replace("•", "• ")
    s = re.sub(r"• +", "• ", s)
    # common PDF Thai glyph splits that still leave readable words
    s = s.replace("รนุ่", "รุ่น").replace("รนุ่", "รุ่น").replace("ร่นุ", "รุ่น")
    s = s.replace("สญั ญา", "สัญญา").replace("สญัญา", "สัญญา")
    s = s.replace("บรกิ าร", "บริการ").replace("บลิ", "บิล")
    s = s.replace("เดอื น", "เดือน").replace("โปรโมชนั", "โปรโมชัน")
    s = s.replace("สว่ นลด", "ส่วนลด").replace("ส่วนลด | จาก | ราคาปกติ", "ส่วนลดจากราคาปกติ")
    s = s.replace("ราคาตอ่ | เดือน", "ราคาต่อเดือน").replace("ราคาต่อ | เดือน", "ราคาต่อเดือน")
    s = s.replace("ทกุ ๆ", "ทุกๆ").replace("ทุกๆ ", "ทุกๆ ")
    s = s.replace("ควิ", "คิว").replace("ลติ ร", "ลิตร")
    s = s.replace("มอนิเตอร ์", "มอนิเตอร์").replace("โทรทศั น์", "โทรทัศน์")
    s = s.replace("No | Service", "No Service")
    s = re.sub(r"\s+", " ", s)
    s = s.replace("| |", "|").strip(" |")
    return s


def parse_money(s: str):
    s = clean_cell(s).replace("•", "").replace("-", "").strip()
    s = s.replace(" ", "")
    if not s or s in {".", "–"}:
        return None
    # 1.149 in AC page is thousands with a dot typo
    if re.fullmatch(r"\d\.\d{3}", s):
        s = s.replace(".", "")
    if re.fullmatch(r"\d{1,3}(?:,\d{3})+(?:\.\d+)?", s) or re.fullmatch(r"\d+(?:\.\d+)?", s):
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    return None


def looks_like_contract(s: str) -> bool:
    t = clean_cell(s).replace(" ", "")
    t = t.replace("5YVisit", "5Y_Visit").replace("5YSelf", "5Y_Self")
    t = t.replace("6YVisit", "6Y_Visit").replace("6YSelf", "6Y_Self")
    t = t.replace("7YVisit", "7Y_Visit").replace("7YSelf", "7Y_Self")
    t = t.replace("2YVisit", "2Y_Visit").replace("2YSelf", "2Y_Self")
    t = re.sub(r"(\dY)(Visit|Self)", r"\1_\2", t)
    return bool(CONTRACT_RE.match(t) or re.match(r"^\d+Y_", t))


def normalize_contract(s: str) -> str:
    t = clean_cell(s)
    t = re.sub(r"\s+", " ", t)
    t = t.replace("No | Service", "No Service")
    t = re.sub(r"(\d+Y)\s+(Visit|Self)", r"\1_\2", t)
    t = t.replace(" ", "") if re.match(r"^\d+Y", t.replace(" ", "")) else t
    t = t.replace("5YVisit", "5Y_Visit").replace("6YVisit", "6Y_Visit")
    t = t.replace("7YVisit", "7Y_Visit").replace("2YVisit", "2Y_Visit")
    t = t.replace("5YSelf", "5Y_Self").replace("6YSelf", "6Y_Self")
    t = t.replace("7YSelf", "7Y_Self").replace("2YSelf", "2Y_Self")
    if t.upper() in {"5Y", "6Y", "7Y"}:
        return t.upper()
    return t


def extract_model_code(text: str) -> str:
    t = clean_cell(text)
    t = t.replace("NEW | ", "NEW ").replace("NEW ", "NEW ")
    low = t.lower()
    if "xboom grab" in low or "lg xboom grab" in low:
        return "LG xboom Grab"
    if "xboom bounce" in low:
        return "LG xboom Bounce"
    if "stage301" in low or "xboom stage" in low:
        return "LG xboom STAGE301"
    # Prefer dotted SKU like WD516AN.ACNPLMT / GC-X257CMHW.AEEPLMT
    dotted = re.findall(r"[A-Z0-9][A-Z0-9-]{2,}\.[A-Z0-9][A-Z0-9._-]+", t, re.I)
    if dotted:
        return dotted[0].rstrip(".")
    m = MODEL_TOKEN_RE.search(t)
    if m:
        return m.group(1).rstrip(".")
    # fallback: first latin token
    m2 = re.search(r"[A-Z][A-Z0-9._-]{3,}", t)
    return (m2.group(0).rstrip(".") if m2 else t[:80])


def split_model(text: str) -> tuple[str, str]:
    t = clean_cell(text)
    code = extract_model_code(t)
    rest = t
    if code:
        rest = t.replace(code, "", 1)
    rest = rest.replace("NEW |", "").replace("NEW", "")
    rest = re.sub(r"^[\s|,:-]+", "", rest).strip(" |")
    return code.strip(), rest


PAGE_CATEGORY = {
    1: "ปก",
    2: "Policy",
    3: "เครื่องกรองน้ำ",
    4: "เครื่องกรองน้ำ",
    5: "เครื่องดูดฝุ่น",
    6: "ตู้เย็น / Plumbing",
    7: "ตู้เย็น / Side-by-Side",
    8: "ตู้เย็น / Side-by-Side",
    9: "ตู้เย็น / Multi-Door",
    10: "ตู้เย็น / Multi-Door",
    11: "ตู้เย็น / Multi-Door",
    12: "ตู้เย็น / 2 ประตู",
    13: "Wash Tower",
    14: "Wash Tower",
    15: "Wash Tower",
    16: "Wash Tower",
    17: "Wash Tower",
    18: "เครื่องซักผ้า",
    19: "เครื่องซักผ้า",
    20: "เครื่องซักผ้า",
    21: "เครื่องซักผ้า / อบผ้า (New Model)",
    22: "เครื่องซักผ้า / อบผ้า (New Model)",
    23: "เครื่องซักผ้า / อบผ้า (New Model)",
    24: "เครื่องล้างจาน",
    25: "เครื่องล้างจาน",
    26: "ตู้ถนอมผ้า (Styler)",
    27: "ไมโครเวฟ",
    28: "เครื่องลดความชื้น",
    29: "เครื่องฟอกอากาศ",
    30: "เครื่องฟอกอากาศ",
    31: "เครื่องฟอกอากาศ",
    32: "เครื่องฟอกอากาศ",
    33: "เครื่องปรับอากาศ RAC (ARTCOOL)",
    34: "เครื่องปรับอากาศ RAC (DUALCOOL SAQ)",
    35: "เครื่องปรับอากาศ RAC (DUALCOOL SAQ)",
    36: "เครื่องปรับอากาศ RAC (DUALCOOL SIQ)",
    37: "เครื่องปรับอากาศ RAC (DUALCOOL SIQ)",
    38: "เครื่องปรับอากาศ SAC (4 Way Cassette)",
    39: "เครื่องปรับอากาศ SAC (1 Way Cassette)",
    40: "เครื่องปรับอากาศ SAC (Round Cassette)",
    41: "โทรทัศน์",
    42: "โทรทัศน์ + Sound bar",
    43: "โทรทัศน์ Lifestyle",
    44: "โทรทัศน์ OLED",
    45: "โทรทัศน์ QNED",
    46: "โทรทัศน์ NANO/UHD",
    47: "โทรทัศน์ Lifestyle",
    48: "มอนิเตอร์",
    49: "มอนิเตอร์ + Grab",
    50: "xboom",
    51: "Sound bar",
    52: "เทียบสเปก Sound bar",
    53: "Combo Promotion",
    54: "เงื่อนไข Combo",
    55: "วิธีคีย์ Combo ใน OSMS",
    56: "วิธีคีย์ Combo ใน OSMS",
    57: "Combo Line-up",
    58: "Combo Line-up",
    59: "Combo Line-up",
    60: "Combo Line-up",
}


def page_category_map(layout_pages: list[str]) -> dict[int, dict]:
    cats = {}
    for i, text in enumerate(layout_pages, 1):
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        title = lines[0] if lines else ""
        joined = " ".join(lines[:12])
        promo = ""
        m = PROMO_HEADER_RE.search(joined)
        if m:
            promo = m.group(1)
        elif "12 ส.ค" in joined:
            promo = "12 ส.ค.-31 ส.ค. 2569 (NEW)"
        elif "11 ส.ค" in joined:
            promo = "11 ส.ค.-31 ส.ค. 2569"
        elif "13 ส.ค" in joined:
            promo = "13 ส.ค.-31 ส.ค. 2569"
        elif "10 ส.ค" in joined:
            promo = "10 ส.ค.-31 ส.ค. 2569"
        elif "4 ส.ค" in joined:
            promo = "4 ส.ค.-31 ส.ค. 2569"
        elif "1 ส.ค" in joined or "1ส.ค" in joined:
            promo = "1 ส.ค.-31 ส.ค. 2569"
        pay = ""
        if "1st Payment 1M" in joined:
            pay = "ชำระล่วงหน้า 1 เดือน (ยกเว้น OUTRIGHT)"
        elif "1st Payment 6M" in joined:
            pay = "ชำระล่วงหน้า 6 เดือน (Advance 50% / 12 บิล)"
        notes = []
        for ln in lines:
            if any(k in ln for k in ["**", "พื้นที่", "ของแถม", "สินค้า", "หมายเหตุ", "ยกเว้น", "รับฟรี"]):
                if len(ln) > 12 and not ln.startswith("ราคา"):
                    notes.append(ln)
        cats[i] = {
            "category": PAGE_CATEGORY.get(i, title),
            "promo_period": promo,
            "payment_note": pay,
            "notes": " | ".join(notes[:3]),
            "raw_first": title,
        }
    return cats


def parse_bill_prices(promo_price: str) -> dict:
    out = {"ราคาโปร_รอบบิล1": None, "ช่วงราคาโปร": "", "ราคาโปร_หลังโปร": None}
    found = BILL_PRICE_RE.findall(clean_cell(promo_price).replace("ท่ี", "ที่").replace("ท ี่", "ที่"))
    parts = []
    for rng, amt in found:
        rng = re.sub(r"\s+", "", rng)
        val = parse_money(amt)
        parts.append(f"{rng}={amt}")
        if rng in {"1", "1-12", "1–12"} or rng.startswith("1-"):
            if out["ราคาโปร_รอบบิล1"] is None:
                out["ราคาโปร_รอบบิล1"] = val
        else:
            out["ราคาโปร_หลังโปร"] = val
    out["ช่วงราคาโปร"] = " | ".join(parts)
    return out


def is_header_row(row: list[str]) -> bool:
    joined = " ".join(row)
    return "แบบการขาย" in joined or "ราคาต่อเดือน" in joined or "รายการสัญญา" in joined


def is_noise_row(row: list[str]) -> bool:
    joined = " ".join(row).strip()
    if not joined:
        return True
    compact = joined.replace(" ", "").replace("|", "")
    if compact in {"NoService", "No|Service"}:
        return True
    if looks_like_contract(joined) and all(not parse_money(c) for c in row):
        # lone "No Service" continuation
        if "No Service" in joined or compact == "NoService":
            return True
    return False


def parse_standard_row(row: list[str], n_cols: int) -> dict | None:
    # 10-col: sale, img, model, contract, cycle, price, promo, promo_price, discount, policy
    # 9-col SAC: sale, img, model, contract, cycle, price, promo, discount, policy
    while len(row) < 10:
        row.append("")
    sale, _img, model, contract, cycle, price, promo, promo_price, discount, policy = row[:10]
    if n_cols == 9 and not looks_like_contract(contract):
        # some 9-col still match 10 layout missing last
        pass
    if n_cols == 9:
        # detect SAC layout: promo often '-' and policy in last col
        # [sale, img, model, contract, cycle, price, promo, discount, policy]
        sale, _img, model, contract, cycle, price, promo, discount, policy = row[:9]
        promo_price = ""
    if not looks_like_contract(contract):
        return None
    price_n = parse_money(price)
    disc_n = parse_money(discount)
    if price_n is None and disc_n is None and not policy:
        return None
    code, detail = split_model(model)
    bills = parse_bill_prices(promo_price)
    return {
        "แบบการขาย": clean_cell(sale),
        "รุ่น": code,
        "รายละเอียดรุ่น": detail,
        "รายการสัญญา": normalize_contract(contract),
        "รอบบริการ": clean_cell(cycle),
        "ราคาปกติต่อเดือน": price_n,
        "เงินล่วงหน้า": None,
        "รายละเอียดโปรโมชัน": clean_cell(promo),
        "ราคาโปร_ข้อความ": clean_cell(promo_price),
        **bills,
        "ส่วนลดจากราคาปกติ": disc_n,
        "Policy": clean_cell(policy).replace(" | ", ""),
    }


def parse_tv_row(row: list[str]) -> dict | None:
    # 9-col: sale, model, contract, advance, monthly, promo, promo_price, discount, policy
    while len(row) < 9:
        row.append("")
    sale, model, contract, advance, monthly, promo, promo_price, discount, policy = row[:9]
    if "No Service" in clean_cell(contract) and not parse_money(monthly) and not parse_money(advance):
        return None
    if not looks_like_contract(contract) and not re.match(r"^\d+Y$", clean_cell(contract).replace(" ", "")):
        return None
    if not parse_money(monthly) and not parse_money(advance):
        return None
    code, detail = split_model(model)
    bills = parse_bill_prices(promo_price)
    return {
        "แบบการขาย": "Subscription",
        "รุ่น": code,
        "รายละเอียดรุ่น": detail,
        "รายการสัญญา": "No Service / " + normalize_contract(contract)
        if "NOSERVICE" in clean_cell(policy).upper()
        else normalize_contract(contract),
        "รอบบริการ": "No Service",
        "ราคาปกติต่อเดือน": parse_money(monthly),
        "เงินล่วงหน้า": parse_money(advance),
        "รายละเอียดโปรโมชัน": clean_cell(promo),
        "ราคาโปร_ข้อความ": clean_cell(promo_price),
        **bills,
        "ส่วนลดจากราคาปกติ": parse_money(discount),
        "Policy": clean_cell(policy).replace(" | ", ""),
    }


def parse_page4(_layout_text: str) -> list[dict]:
    """Page 4 table is drawn as overlapping text boxes; values are taken from the rendered page."""
    wd_detail = (
        "WD516AN.ACNPLMT (น้ำเงิน) / WD516AN.AEWPLMT (ขาว) / WD516AN.ASLPLMT (เงิน) / "
        "WD518AN.ABGPLMT (เบจ) / WD518AN.AWHPLMT (ขาว) / WD518AN.ACGPLMT (เทา)"
    )
    promo = "ราคา 149.- (1 รอบบิลแรก) | ลด 50% 11 เดือน (รอบบิลที่ 2-12)"
    specs = [
        ("WD516AN / WD518AN", wd_detail, "5Y_Visit", 799, "1=149 | 2-12=399 | 13-60=799", 5050, "VISIT_5Y_6M_PRO_149(1M)_DC50%(11M)"),
        ("WD516AN / WD518AN", wd_detail, "5Y_Self", 699, "1=149 | 2-12=349 | 13-60=699", 4400, "SELF_5Y_6M_PRO_149(1M)_DC50%(11M)"),
        ("WD516AN / WD518AN", wd_detail, "7Y_Visit", 599, "1=149 | 2-12=299 | 13-84=599", 3750, "VISIT_7Y_6M_PRO_149(1M)_DC50%(11M)"),
        ("WD516AN / WD518AN", wd_detail, "7Y_Self", 499, "1=149 | 2-12=249 | 13-84=499", 3100, "SELF_7Y_6M_PRO_149(1M)_DC50%(11M)"),
        ("WD110MN.ABGPLMT", "Calming Beige / 132 x 358 x 230 มม.", "5Y_Visit", 549, "1=149 | 2-12=274 | 13-60=549", 3425, "VISIT_5Y_6M_PRO_149(1M)_DC50%(11M)"),
        ("WD110MN.ABGPLMT", "Calming Beige / 132 x 358 x 230 มม.", "5Y_Self", 499, "1=149 | 2-12=249 | 13-60=499", 3100, "SELF_5Y_6M_PRO_149(1M)_DC50%(11M)"),
        ("WD110MN.ABGPLMT", "Calming Beige / 132 x 358 x 230 มม.", "7Y_Visit", 449, "1=149 | 2-12=224 | 13-84=449", 2775, "VISIT_7Y_6M_PRO_149(1M)_DC50%(11M)"),
        ("WD110MN.ABGPLMT", "Calming Beige / 132 x 358 x 230 มม.", "7Y_Self", 399, "1=149 | 2-12=199 | 13-84=399", 2450, "SELF_7Y_6M_PRO_149(1M)_DC50%(11M)"),
    ]
    rows = []
    for code, detail, contract, price, bills, disc, policy in specs:
        bill_txt = " ".join(f"รอบบิลที่ {a} ({b})" for a, b in (p.split("=") for p in bills.split(" | ")))
        rows.append(
            {
                "แบบการขาย": "Subscription",
                "รุ่น": code,
                "รายละเอียดรุ่น": detail,
                "รายการสัญญา": contract,
                "รอบบริการ": "ทุกๆ 6 เดือน",
                "ราคาปกติต่อเดือน": float(price),
                "เงินล่วงหน้า": None,
                "รายละเอียดโปรโมชัน": promo,
                "ราคาโปร_ข้อความ": bills,
                **parse_bill_prices(bill_txt),
                "ส่วนลดจากราคาปกติ": float(disc),
                "Policy": policy,
            }
        )
    return rows


def extract_rows(pdf_path: Path, cats: dict) -> list[dict]:
    all_rows: list[dict] = []
    with pdfplumber.open(str(pdf_path)) as doc:
        for i, pg in enumerate(doc.pages, 1):
            meta = cats.get(i, {})
            tables = pg.extract_tables() or []
            page_rows = []
            if i == 4:
                layout = LAYOUT_TXT.read_text(encoding="utf-8", errors="replace").split("\x0c")
                page_rows = parse_page4(layout[3] if len(layout) > 3 else "")
            else:
                good = [t for t in tables if t and len(t) >= 3]
                for table in good:
                    sale_ff = ""
                    model_ff = ""
                    n_cols = max(len(r) for r in table)
                    tv_like = i >= 41
                    for raw in table:
                        row = [clean_cell(c) for c in raw]
                        if is_header_row(row) or is_noise_row(row):
                            continue
                        # forward fill
                        if row[0] in {"Subscription", "Outright", "Subscribe"}:
                            sale_ff = row[0]
                        elif sale_ff and not row[0]:
                            row[0] = sale_ff
                        model_idx = 1 if tv_like else 2
                        if model_idx < len(row) and extract_model_code(row[model_idx]) and len(extract_model_code(row[model_idx])) >= 5:
                            model_ff = row[model_idx]
                        elif model_ff and model_idx < len(row) and not row[model_idx]:
                            row[model_idx] = model_ff
                        parsed = parse_tv_row(row) if tv_like else parse_standard_row(row, n_cols)
                        if not parsed:
                            continue
                        if not parsed["แบบการขาย"]:
                            parsed["แบบการขาย"] = sale_ff or "Subscription"
                        if not parsed["รุ่น"] and model_ff:
                            parsed["รุ่น"], parsed["รายละเอียดรุ่น"] = split_model(model_ff)
                        page_rows.append(parsed)
            for r in page_rows:
                r["หน้า"] = i
                r["หมวดสินค้า"] = meta.get("category", "")
                r["ระยะเวลาโปรโมชัน"] = meta.get("promo_period", "")
                r["เงื่อนไขชำระ"] = meta.get("payment_note", "")
                r["หมายเหตุหน้า"] = meta.get("notes", "")
                all_rows.append(r)
    return all_rows


def policy_rows() -> list[dict]:
    return [
        {
            "Policy ตัวอย่าง": "VISIT_6Y_24M",
            "ประเภท": "VISIT",
            "ความหมาย": "ผู้เชี่ยวชาญจากแอลจีเข้าให้บริการถึงบ้าน เมื่อถึงรอบบริการ ช่างศูนย์บริการแอลจีนัดหมายเข้าบำรุงรักษาและเปลี่ยนอะไหล่ถึงบ้าน",
            "อายุสัญญา": "6 ปี (72 รอบบิล)",
            "รอบบริการ": "ทุก 24 เดือน",
        },
        {
            "Policy ตัวอย่าง": "SELF_5Y_6M",
            "ประเภท": "SELF",
            "ความหมาย": "ลูกค้าเปลี่ยนอะไหล่ด้วยตนเอง ไม่มีช่างเข้าบ้าน เมื่อถึงรอบบริการจะส่งอะไหล่ทางพัสดุให้ลูกค้าเปลี่ยนเอง",
            "อายุสัญญา": "5 ปี (60 รอบบิล)",
            "รอบบริการ": "ส่งอะไหล่ทุก 6 เดือน",
        },
        {
            "Policy ตัวอย่าง": "NOSERVICE_5Y",
            "ประเภท": "No Service",
            "ความหมาย": "ไม่จัดส่งอะไหล่/ไม่เข้าบริการที่บ้าน แต่สินค้ายังได้การรับประกันตลอดอายุสัญญา",
            "อายุสัญญา": "5 ปี (60 รอบบิล)",
            "รอบบริการ": "ไม่มี",
        },
        {
            "Policy ตัวอย่าง": "On site Service",
            "ประเภท": "การรับประกัน",
            "ความหมาย": "ช่างเข้าซ่อมถึงบ้าน: สินค้า LG Subscribe ทุกประเภท รวมหุ่นยนต์ดูดฝุ่นเฉพาะรุ่น A9T-ULTRA (ลูกค้าโทรศูนย์บริการแอลจีเพื่อนัดหมาย)",
            "อายุสัญญา": "",
            "รอบบริการ": "",
        },
        {
            "Policy ตัวอย่าง": "Carry-in",
            "ประเภท": "การรับประกัน",
            "ความหมาย": "ลูกค้านำส่งซ่อมที่ศูนย์บริการแอลจี: ไมโครเวฟ, ซาวด์บาร์, xboom Bluetooth speaker, หุ่นยนต์ดูดฝุ่น A9T-CORE / A9T-LITE",
            "อายุสัญญา": "",
            "รอบบริการ": "",
        },
        {
            "Policy ตัวอย่าง": "สินค้า No Service",
            "ประเภท": "รายการสินค้า",
            "ความหมาย": "1. โทรทัศน์ 2. มอนิเตอร์ 3. ไมโครเวฟ 4. ซาวด์บาร์ 5. xboom GRAB 6. xboom BOUNCE 7. xboom STAGE301",
            "อายุสัญญา": "",
            "รอบบริการ": "",
        },
    ]


def combo_rows() -> list[dict]:
    return [
        {
            "หัวข้อ": "โปรโมชันคอมโบ ส.ค. 2569",
            "รายละเอียด": "ลด 10% ทุกรอบบิล ระหว่าง 1-31 ส.ค. 2569 หรือจนกว่าจะมีประกาศเปลี่ยน",
        },
        {
            "หัวข้อ": "Exist Customer",
            "รายละเอียด": "ลูกค้าเก่าที่เคยซื้อ LG Subscribe แล้ว (สัญญายังอยู่หรือสิ้นสุดแล้ว) ซื้อเพิ่มตั้งแต่ 1 เครื่องขึ้นไป ลด 10% — Promotion: COMBO_SPECIAL_ADD_(1 UNIT UP) เลือกใน OSMS แบบ EXIST+NEW",
        },
        {
            "หัวข้อ": "New Customer",
            "รายละเอียด": "ลูกค้าที่ไม่เคยซื้อ LG Subscribe รวมถึงสั่งแล้วแต่ยังไม่ติดตั้ง ซื้อตั้งแต่ 2 เครื่องขึ้นไป ลด 10% — Promotion: COMBO_NEW_SPECIAL_(2 UNITS UP) เลือกใน OSMS แบบ NEW+NEW",
        },
        {
            "หัวข้อ": "ที่อยู่ติดตั้ง Exist",
            "รายละเอียด": "ไม่จำเป็นต้องเป็นที่อยู่เดิม แต่ถ้า 1 ออเดอร์มีหลายสัญญา ต้องติดตั้งที่อยู่เดียวกัน ยกเว้นพิสูจน์ได้ว่าเป็นลูกค้าคนเดียวกัน/เครือญาติ",
        },
        {
            "หัวข้อ": "ที่อยู่ติดตั้ง New",
            "รายละเอียด": "ที่อยู่ติดตั้งทุกสินค้าต้องเป็นที่อยู่เดียวกัน",
        },
        {
            "หัวข้อ": "ยกเลิกบางรายการก่อนติดตั้ง",
            "รายละเอียด": "ถ้าลูกค้ายกเลิกสินค้าบางชิ้นในออเดอร์คอมโบ ออเดอร์ที่เหลือจะถูกยกเลิกทั้งหมด ต้องสั่งซื้อใหม่",
        },
        {
            "หัวข้อ": "เครดิตไม่ผ่านบางรุ่น",
            "รายละเอียด": "ต้องคีย์ใหม่ทั้งออเดอร์ เพราะส่วนลดเปลี่ยนตามจำนวนสินค้า",
        },
        {
            "หัวข้อ": "คืนสินค้าภายใน 7 วันหลังติดตั้ง",
            "รายละเอียด": "ส่วนลดรายเดือนจะถูกปรับในระบบอัตโนมัติ",
        },
        {
            "หัวข้อ": "ใช้ได้กับ",
            "รายละเอียด": "เฉพาะประเภท Subscribe ใช้กับ Outright ไม่ได้",
        },
        {
            "หัวข้อ": "หน้า 52",
            "รายละเอียด": "ตารางเทียบสเปก Sound Bar: S95TR 810W 9.1.5Ch / S80TY 480W 3.1.3Ch / S70TY 400W 3.1.1Ch / S30A 150W 2.1Ch",
        },
        {
            "หัวข้อ": "หน้า 57-60",
            "รายละเอียด": "ภาพ Line-up รุ่นที่เข้า Combo (WP, WM, REF, AP, RAC/SAC, TV, AV, Monitor, Microwave) ไม่มีตัวเลขราคา — เช็คราคาที่ชีทรายการราคา",
        },
    ]


def style_sheet(ws, header_fill="A50034"):
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Calibri", size=11)
    fill = PatternFill("solid", fgColor=header_fill)
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32
    # widths
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        maxlen = 0
        for cell in col[:80]:
            val = "" if cell.value is None else str(cell.value)
            maxlen = max(maxlen, min(len(val), 48))
        ws.column_dimensions[letter].width = max(12, min(maxlen + 4, 42))
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"


def write_excel(rows: list[dict], cats: dict):
    df = pd.DataFrame(rows)
    col_order = [
        "หน้า",
        "หมวดสินค้า",
        "แบบการขาย",
        "รุ่น",
        "รายละเอียดรุ่น",
        "รายการสัญญา",
        "รอบบริการ",
        "ราคาปกติต่อเดือน",
        "เงินล่วงหน้า",
        "ระยะเวลาโปรโมชัน",
        "รายละเอียดโปรโมชัน",
        "ราคาโปร_รอบบิล1",
        "ช่วงราคาโปร",
        "ส่วนลดจากราคาปกติ",
        "Policy",
        "เงื่อนไขชำระ",
        "หมายเหตุหน้า",
        "ราคาโปร_ข้อความ",
    ]
    for c in col_order:
        if c not in df.columns:
            df[c] = None
    df = df[col_order]
    df.insert(0, "เช็คแล้ว", "")
    df.insert(1, "ผลเช็ค", "")
    df.insert(2, "หมายเหตุตอนเช็ค", "")

    # derived flags
    def flag(r):
        notes = []
        if pd.isna(r["ราคาปกติต่อเดือน"]):
            notes.append("ไม่มีราคาปกติ")
        if pd.isna(r["ส่วนลดจากราคาปกติ"]) and r["แบบการขาย"] != "Outright":
            if r.get("รายละเอียดโปรโมชัน") not in {"-", ""} and pd.notna(r.get("รายละเอียดโปรโมชัน")):
                if "149" in str(r.get("รายละเอียดโปรโมชัน", "")):
                    pass
        if r["Policy"] in {"", None}:
            notes.append("ไม่มี Policy")
        return " | ".join(notes)

    df["จุดที่ควรเช็ค"] = df.apply(flag, axis=1)

    summary = (
        df.groupby(["หมวดสินค้า", "แบบการขาย"], dropna=False)
        .agg(
            จำนวนรายการ=("รุ่น", "count"),
            จำนวนรุ่น=("รุ่น", "nunique"),
            ราคาปกติต่ำสุด=("ราคาปกติต่อเดือน", "min"),
            ราคาปกติสูงสุด=("ราคาปกติต่อเดือน", "max"),
        )
        .reset_index()
    )

    howto = pd.DataFrame(
        [
            {"ลำดับ": 1, "วิธีเช็ค": "เปิดชีท 'รายการราคา' แล้วใช้ Filter ที่แถวหัวตาราง"},
            {"ลำดับ": 2, "วิธีเช็ค": "กรอง 'หมวดสินค้า' ตามหน้าที่กำลังดูใน PDF เช่น เครื่องกรองน้ำ, Wash Tower, โทรทัศน์"},
            {"ลำดับ": 3, "วิธีเช็ค": "เทียบคอลัมน์ รุ่น + รายการสัญญา + ราคาปกติต่อเดือน + Policy กับหน้า PDF"},
            {"ลำดับ": 4, "วิธีเช็ค": "คอลัมน์ ช่วงราคาโปร รวมรอบบิลที่ 1 / รอบลด / รอบปกติ เช่น 1=149 | 2-12=399 | 13-60=799"},
            {"ลำดับ": 5, "วิธีเช็ค": "ใส่เครื่องหมายใน 'เช็คแล้ว' และเลือกผลเช็ค ตรง / ไม่ตรง / ต้องถามเพิ่ม"},
            {"ลำดับ": 6, "วิธีเช็ค": "ถ้าไม่ตรง ใส่หน้าที่ PDF และค่าที่เห็นจริงใน 'หมายเหตุตอนเช็ค'"},
            {"ลำดับ": 7, "วิธีเช็ค": "หน้า 1 = ปกโปร 1-31 ส.ค. 2569, หน้า 2 = ความหมาย VISIT/SELF/No Service"},
            {"ลำดับ": 8, "วิธีเช็ค": "หน้า 52-60 ไม่ใช่ตารางราคา แต่เป็นสเปกซาวด์บาร์ / เงื่อนไขคอมโบ / วิธีคีย์ OSMS / รูป line-up"},
            {"ลำดับ": 9, "วิธีเช็ค": "เอกสารต้นทางระบุว่าราคาและระยะเวลาอาจเปลี่ยนตามประกาศบริษัท (ลงวันที่ 13 ส.ค. 2026)"},
            {
                "ลำดับ": 10,
                "วิธีเช็ค": "ตัวอักษรไทยบางช่องอาจเพี้ยนเล็กน้อยเพราะฟอนต์ใน PDF — ให้ยึดรุ่น, ตัวเลขราคา และรหัส Policy เป็นหลัก",
            },
        ]
    )

    cover = pd.DataFrame(
        [
            {"รายการ": "ไฟล์ต้นทาง", "ค่า": "Price list_Aug_V3.pdf"},
            {"รายการ": "เวอร์ชันเอกสาร", "ค่า": "LGETH | 2026.08.13"},
            {"รายการ": "ช่วงโปรหลัก", "ค่า": "1-31 ส.ค. 2569 (บางรุ่นเริ่ม 4 / 10 / 11 / 12 / 13 ส.ค.)"},
            {"รายการ": "จำนวนหน้า PDF", "ค่า": 60},
            {"รายการ": "จำนวนแถวราคาที่ดึงได้", "ค่า": len(df)},
            {"รายการ": "จำนวนรุ่นไม่ซ้ำ", "ค่า": int(df["รุ่น"].nunique())},
            {"รายการ": "วัตถุประสงค์ไฟล์นี้", "ค่า": "ให้เช็คราคา/โปร/Policy เทียบกับ PDF ไม่ใช่ไฟล์สำหรับขึ้นระบบโดยตรง"},
        ]
    )

    notes_pages = []
    for i, meta in cats.items():
        if meta.get("notes"):
            notes_pages.append(
                {
                    "หน้า": i,
                    "หมวด": meta.get("category"),
                    "หมายเหตุจากหน้า": meta.get("notes"),
                    "ระยะเวลาโปร": meta.get("promo_period"),
                    "เงื่อนไขชำระ": meta.get("payment_note"),
                }
            )
    notes_df = pd.DataFrame(notes_pages)

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        cover.to_excel(writer, index=False, sheet_name="ภาพรวม")
        howto.to_excel(writer, index=False, sheet_name="วิธีเช็ค")
        df.to_excel(writer, index=False, sheet_name="รายการราคา")
        summary.to_excel(writer, index=False, sheet_name="สรุปตามหมวด")
        pd.DataFrame(policy_rows()).to_excel(writer, index=False, sheet_name="Policy")
        pd.DataFrame(combo_rows()).to_excel(writer, index=False, sheet_name="Combo")
        notes_df.to_excel(writer, index=False, sheet_name="หมายเหตุท้ายหน้า")

    wb = load_workbook(OUT_PATH)
    fills = {
        "ภาพรวม": "A50034",
        "วิธีเช็ค": "333333",
        "รายการราคา": "A50034",
        "สรุปตามหมวด": "1F4E79",
        "Policy": "7A001F",
        "Combo": "C00000",
        "หมายเหตุท้ายหน้า": "833C0C",
    }
    for name, color in fills.items():
        style_sheet(wb[name], color)

    ws = wb["รายการราคา"]
    # number formats
    headers = {cell.value: cell.column for cell in ws[1]}
    money_cols = ["ราคาปกติต่อเดือน", "เงินล่วงหน้า", "ราคาโปร_รอบบิล1", "ส่วนลดจากราคาปกติ"]
    for h in money_cols:
        if h in headers:
            col = headers[h]
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                cell.number_format = "#,##0.00"

    # dropdown for ผลเช็ค
    if "ผลเช็ค" in headers:
        dv = DataValidation(type="list", formula1='"ตรง,ไม่ตรง,ต้องถามเพิ่ม,ข้าม"', allow_blank=True)
        dv.error = "เลือกจากรายการ"
        dv.prompt = "เลือกผลการเช็ค"
        letter = get_column_letter(headers["ผลเช็ค"])
        dv.add(f"{letter}2:{letter}{ws.max_row}")
        ws.add_data_validation(dv)
    if "เช็คแล้ว" in headers:
        dv2 = DataValidation(type="list", formula1='"Y,"', allow_blank=True)
        letter = get_column_letter(headers["เช็คแล้ว"])
        dv2.add(f"{letter}2:{letter}{ws.max_row}")
        ws.add_data_validation(dv2)

    # highlight empty price
    if "ราคาปกติต่อเดือน" in headers:
        col = get_column_letter(headers["ราคาปกติต่อเดือน"])
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            FormulaRule(formula=[f'AND({col}2="",A2<>"")'], fill=PatternFill("solid", fgColor="F4B183")),
        )

    ws.sheet_properties.tabColor = "A50034"
    wb["ภาพรวม"].sheet_view.showGridLines = False
    wb.save(OUT_PATH)


def main():
    layout_text = LAYOUT_TXT.read_text(encoding="utf-8", errors="replace")
    layout_pages = layout_text.split("\x0c")
    cats = page_category_map(layout_pages)
    rows = extract_rows(PDF_PATH, cats)
    if not rows:
        raise SystemExit("no rows extracted")
    write_excel(rows, cats)
    print(f"rows={len(rows)} models={len({r['รุ่น'] for r in rows})} out={OUT_PATH}")
    # print category counts
    from collections import Counter

    c = Counter((r["หน้า"], r["หมวดสินค้า"]) for r in rows)
    for k, n in sorted(c.items()):
        print(f"  p{k[0]:02d} {k[1]}: {n}")


if __name__ == "__main__":
    main()
